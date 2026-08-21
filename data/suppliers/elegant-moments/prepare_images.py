#!/usr/bin/env python3
"""Recompress Elegant Moments catalog images for upload to Shopify.

The supplier ships ~1200x1800 JPEGs at very low compression (~0.6 MB each,
811 MB for the 2026 collection). Re-encoding at quality 82 keeps the same pixel
dimensions but cuts the payload roughly 5x, which matters because every image is
uploaded over the Admin API. Shopify re-encodes and serves responsive derivatives
from its CDN regardless, so nothing is lost in delivered quality.

Dimensions are never changed — the supplier licence forbids modifying images in a
way that distorts the original proportions.

Usage:
    python prepare_images.py --src "F:/CODING/BikinniSite/Files_Site/Items"
"""

import argparse
import sys
from pathlib import Path

try:
    from PIL import Image
except ImportError:
    sys.exit("Pillow is required: python -m pip install Pillow")

import openpyxl

# Share the workbook parsing with the importer so a header quirk only has to be
# handled once — the Hosiery sheet ships blank headers for its image columns.
from build_import import DESC_SHEETS, repair_header

HERE = Path(__file__).resolve().parent
SOURCE = HERE / "source"
DEST = HERE / "out" / "images"
DESCRIPTIONS = SOURCE / "2026_Collection_Descriptions.xlsx"


def referenced_images():
    """Image filenames the descriptions workbook actually points at."""
    wb = openpyxl.load_workbook(DESCRIPTIONS, read_only=True, data_only=True)
    names = set()
    for sheet in DESC_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(it)]
        header = repair_header(header, sheet)
        for raw in it:
            row = dict(zip(header, raw))
            if not row.get("Style"):
                continue
            for key in ("Image 1", "Image 2", "Image 3", "Image 4"):
                val = row.get(key)
                if val:
                    names.add(str(val).strip().lower())
    return names


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--src", required=True,
                    help="directory of extracted supplier JPEGs")
    ap.add_argument("--quality", type=int, default=82)
    ap.add_argument("--dest", default=str(DEST))
    ap.add_argument("--all", action="store_true",
                    help="convert every JPEG, not just those the workbook references")
    args = ap.parse_args()

    src = Path(args.src)
    if not src.is_dir():
        sys.exit(f"Not a directory: {src}")
    dest = Path(args.dest)
    dest.mkdir(parents=True, exist_ok=True)

    wanted = None if args.all else referenced_images()
    on_disk = {p.name.lower(): p for p in src.glob("*.jpg")}

    if wanted is not None:
        missing = sorted(wanted - set(on_disk))
        if missing:
            print(f"WARNING: {len(missing)} referenced image(s) not in {src}:")
            for name in missing[:10]:
                print(f"  {name}")

    targets = sorted(on_disk) if wanted is None else sorted(wanted & set(on_disk))

    orig_bytes = new_bytes = 0
    for i, name in enumerate(targets, 1):
        srcp = on_disk[name]
        outp = dest / name
        orig_bytes += srcp.stat().st_size
        with Image.open(srcp) as im:
            im = im.convert("RGB")
            im.save(outp, "JPEG", quality=args.quality, optimize=True,
                    progressive=True)
        new_bytes += outp.stat().st_size
        if i % 100 == 0:
            print(f"  {i}/{len(targets)}...")

    saved = 100 * (1 - new_bytes / orig_bytes) if orig_bytes else 0
    print(f"\nConverted {len(targets)} images")
    print(f"  {orig_bytes/1e6:.0f} MB -> {new_bytes/1e6:.0f} MB  ({saved:.0f}% smaller)")
    print(f"  written to {dest}")


if __name__ == "__main__":
    main()
