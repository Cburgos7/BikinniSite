#!/usr/bin/env python3
"""Build a Shopify product import CSV from Elegant Moments wholesale exports.

Joins the live inventory feed (stock + wholesale cost) against the collection
descriptions workbook (copy, fabric, categories, image filenames) on STYLE, keeps
only sellable stock, and emits Shopify's product import format.

Each supplier release is a separate catalogue (see CATALOGUES): same shape,
different sheet names, size column spelling and image column count.

Usage:
    python build_import.py                      # the 2026 collection
    python build_import.py --catalogue vivace   # 2026-2027 Vivace
    python build_import.py --markup 2.5 --image-base https://cdn.example.com/em/

Requires: openpyxl (plus xlrd for the legacy .xls hosiery workbook)
"""

import argparse
import csv
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: python -m pip install openpyxl")

HERE = Path(__file__).resolve().parent
SOURCE = HERE / "source"
OUT = HERE / "out"

INVENTORY = SOURCE / "liveinventory.csv"

# --- Catalogue configuration -------------------------------------------------
#
# Every Elegant Moments descriptions workbook has the same shape but not the
# same spelling: sheet names differ, the size column is "Sizes" in one release
# and "Size" in the next, and the number of image columns grows when a
# collection ships editorial shots. Adding a catalogue is configuration here,
# not a new code path.
#
# Keys:
#   workbook     filename inside source/
#   sheets       ordered {sheet name: tag applied to its products}
#   images       how many "Image N" columns the layout defines; only used to
#                repair blank headers — actual columns are read from the header
#   plus_suffix  style-number suffixes marking a plus-size twin of a parent
#                style. 2026 uses "X" (11022 / 11022X); Vivace and Holiday use
#                "Q" for Queen (82579 / 82579Q).
#   out          output directory
#   follows      catalogues imported earlier whose styles this one must not
#                re-list. Workbooks overlap — 14 of Holiday's styles are already
#                carried by the 2026 and Vivace releases, and importing them
#                twice would collide on the em-<style> handle.

BASE_COLUMNS = [
    "Style", "Description", "Size", "Color", "Page", "Price", "UPC Code",
    "SKU Number", "Country Of Origin", "Fabric/Material", "Category 1",
    "Category 2", "Category 3", "Shown With", "Weight (oz)",
]

# Spelling drift between releases. Normalising on read means the rest of the
# builder only ever sees one name per concept.
COLUMN_ALIASES = {
    "sizes": "Size",
    "pdf page": "Page",
}

CATALOGUES = OrderedDict([
    ("2026", {
        "workbook": "2026_Collection_Descriptions.xlsx",
        "sheets": OrderedDict([
            ("Lingerie", "Lingerie"),
            ("Leather", "Leather"),
            ("Vinyl", "Vinyl"),
            ("Costumes", "Costumes"),
            # The sheet is named "Hosiery Items"; the shopper-facing tag is not.
            ("Hosiery Items", "Hosiery"),
        ]),
        "images": 4,
        "plus_suffix": ("X",),
        "out": OUT,
        "follows": [],
    }),
    ("vivace", {
        "workbook": "2026-2027_Vivace_Descriptions.xlsx",
        "sheets": OrderedDict([
            ("2026-2027 Vivace Swimwear", "Swimwear"),
            ("2026-2027 Vivace", "Vivace"),
            ("2026-2027 Vivace Panties", "Vivace"),
            ("2026-2027 Vivace Menswear", "Menswear"),
        ]),
        # Swim styles carry editorial shots on top of front/back, so this
        # layout runs to Image 8 where the 2026 one stopped at 4.
        "images": 8,
        "plus_suffix": ("Q",),
        "out": OUT / "vivace",
        "follows": ["2026"],
    }),
    ("holiday", {
        "workbook": "2026_Holiday_Descriptions.xlsx",
        "sheets": OrderedDict([("2026 Holiday", "Holiday")]),
        "images": 4,
        "plus_suffix": ("Q", "X"),
        "out": OUT / "holiday",
        "follows": ["2026", "vivace"],
    }),
    ("hosiery", {
        # Legacy .xls — openpyxl cannot read it, so this one goes through xlrd.
        "workbook": "2025-2026_Hosiery_Descriptions.xls",
        "sheets": OrderedDict([("2025-2026 Hosiery", "Hosiery")]),
        "images": 3,
        "plus_suffix": ("Q", "X"),
        "out": OUT / "hosiery",
        "follows": ["2026", "vivace", "holiday"],
    }),
])

DEFAULT_CATALOGUE = "2026"


def canonical_columns(image_count):
    return BASE_COLUMNS + [f"Image {i}" for i in range(1, image_count + 1)]


def normalise_header(header):
    """Map a release's column spellings onto the canonical ones."""
    out = []
    for name in header:
        out.append(COLUMN_ALIASES.get(name.strip().lower(), name))
    return out


IMAGE_COLUMN_RE = re.compile(r"^image\s*(\d+)$", re.I)


def image_columns(header):
    """Image columns present in a header, in numeric order.

    Reading these from the header rather than a fixed list is what lets a
    workbook with Image 1..8 keep its editorial shots instead of losing
    everything past Image 4.
    """
    found = []
    for name in header:
        m = IMAGE_COLUMN_RE.match(str(name).strip())
        if m:
            found.append((int(m.group(1)), name))
    return [name for _, name in sorted(found)]


def repair_header(header, sheet_name, canonical=None, quiet=False):
    """Fill in blank header cells from the canonical column order.

    Only fills positions that are blank, and only where the named columns that
    ARE present still line up with the canonical order — otherwise a genuinely
    different layout would get mislabelled.

    The 2026 "Hosiery Items" sheet leaves columns 13-18 unnamed even though the
    rows carry the same data as every other sheet — without this, weights and
    image filenames are silently dropped. Later workbooks name every column that
    holds data but pad the sheet with a few trailing blank ones; those are left
    alone because `canonical` stops where the data does.
    """
    if canonical is None:
        canonical = canonical_columns(4)
    for i, name in enumerate(header):
        if name and i < len(canonical) and name != canonical[i]:
            if not quiet:
                print(f"  WARNING: {sheet_name!r} column {i} is {name!r}, "
                      f"expected {canonical[i]!r} — leaving header untouched")
            return header

    # Only blanks inside the header are filled. Nothing is appended past its
    # end: a sheet whose header is shorter than the canonical list simply has
    # fewer columns, and inventing names for columns that do not exist would
    # report image columns the workbook never shipped.
    repaired = list(header)
    filled = []
    for i, name in enumerate(canonical[:len(repaired)]):
        if not repaired[i]:
            repaired[i] = name
            filled.append(name)
    if filled and not quiet:
        print(f"  note: {sheet_name!r} had {len(filled)} blank header(s); "
              f"filled positionally: {', '.join(filled)}")
    return repaired


def sheet_rows(path, sheet_name):
    """Yield (header, row tuples) for one sheet of an .xlsx or legacy .xls.

    The 2025-2026 Hosiery workbook is still in the pre-2007 binary format that
    openpyxl refuses outright, so it goes through xlrd. xlrd hands every number
    back as a float, which would turn style 12179 into "12179.0" and break the
    join against the inventory feed — integral floats are folded back to int.
    """
    if path.suffix.lower() == ".xls":
        try:
            import xlrd
        except ImportError:
            sys.exit(f"{path.name} is a legacy .xls; xlrd is required to read "
                     f"it: python -m pip install xlrd")
        book = xlrd.open_workbook(path)
        if sheet_name not in book.sheet_names():
            return None, []
        sheet = book.sheet_by_name(sheet_name)

        def value(cell):
            if cell.ctype == xlrd.XL_CELL_NUMBER and float(cell.value).is_integer():
                return int(cell.value)
            return cell.value if cell.ctype != xlrd.XL_CELL_EMPTY else None

        rows = [tuple(value(c) for c in sheet.row(i))
                for i in range(sheet.nrows)]
        if not rows:
            return None, []
        return rows[0], rows[1:]

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        return None, []
    it = wb[sheet_name].iter_rows(values_only=True)
    try:
        header = next(it)
    except StopIteration:
        return None, []
    return header, list(it)


def read_sheet(path, sheet_name, image_count, quiet=False):
    """Return (header, rows) with the header normalised and repaired."""
    raw_header, rows = sheet_rows(path, sheet_name)
    if raw_header is None:
        return None, []
    header = [str(h).strip() if h is not None else "" for h in raw_header]
    header = normalise_header(header)
    header = repair_header(header, sheet_name, canonical_columns(image_count),
                           quiet=quiet)
    return header, rows

# The supplier licence requires their name to appear in advertising content.
ATTRIBUTION = "Elegant Moments"

# Shopify sorts variant options in the order it first sees them, so emit sizes in
# wearable order rather than the alphabetical order the source happens to use.
SIZE_ORDER = [
    "XS", "S", "S/M", "M", "M/L", "L", "L/XL", "XL",
    "1X", "1X/2X", "2X", "3X", "3X/4X", "4X", "5X",
    "O/S", "OS", "Q/S", "Q'S",
]
SIZE_RANK = {s: i for i, s in enumerate(SIZE_ORDER)}

# Numeric sizes (bra bands: 32, 34, ...) sort after named sizes, numerically.
NUMERIC_SIZE_BASE = len(SIZE_ORDER)

SHOPIFY_COLUMNS = [
    "Handle", "Title", "Body (HTML)", "Vendor", "Type", "Tags", "Published",
    "Option1 Name", "Option1 Value", "Option2 Name", "Option2 Value",
    "Variant SKU", "Variant Grams", "Variant Inventory Tracker",
    "Variant Inventory Qty", "Variant Inventory Policy",
    "Variant Fulfillment Service", "Variant Price", "Variant Compare At Price",
    "Variant Requires Shipping", "Variant Taxable", "Variant Barcode",
    "Image Src", "Image Position", "Image Alt Text", "Status",
]


def size_sort_key(size):
    s = (size or "").strip().upper()
    if s in SIZE_RANK:
        return (SIZE_RANK[s], 0)
    if s.isdigit():
        return (NUMERIC_SIZE_BASE, int(s))
    return (NUMERIC_SIZE_BASE + 1, 0)


def slugify(value):
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug or "item"


def strip_trade_notes(text):
    """Remove wholesale-facing asides from copy meant for shoppers.

    The supplier's descriptions carry notes aimed at buyers, not customers —
    "*Available Boxed" is about how the item ships to a retailer, and it reads as
    noise on a product page.
    """
    text = re.sub(r'\*\s*Available\s+Boxed\.?', '', text, flags=re.I)
    text = re.sub(r'\s{2,}', ' ', text)
    return text.strip()


def titleize(text, style):
    """Build a product title from the description's leading clause.

    Descriptions read as sentences ("Eyelash lace and satin babydoll with
    underwire cups, adjustable straps and hook"). The clause before the first
    comma / "with" is the product name; the rest is detail.
    """
    text = strip_trade_notes(text or "")
    if not text:
        return f"Style {style}"
    # Costume rows use "Name - description"; keep the name.
    if " - " in text[:60]:
        head = text.split(" - ", 1)[0]
    else:
        head = re.split(r",| with | w/ ", text, maxsplit=1)[0]
    head = head.strip(" .")
    words = head.split()
    if len(words) > 10:
        head = " ".join(words[:10])
    return head[:1].upper() + head[1:] if head else f"Style {style}"


# An order has to cover costs the item itself never sees: the supplier's $3.50
# per-ORDER drop-ship fee, their $4-12 postage (unquotable until the parcel is
# packed), and payment processing. On top of that the store runs an influencer
# programme paying 15% commission on orders carrying a 10%-off code.
#
# The customer discount is the expensive half, not the commission: it cuts revenue
# AND the base the commission is calculated on. At 2.75x with no discount a
# floor-priced item affords 15.7% commission; add 10% off and that collapses to
# 6.7%. So markup and floor were raised together -- neither lever is enough alone.
#
# At 2.75x with a $21.95 floor, the worst case in the catalogue (cheapest item,
# dearest postage, discount and commission both applied) still affords 15.8%
# commission, so the programme pays on EVERY order rather than only above a
# minimum basket. See .planning/research/REFERRALS.md.
DEFAULT_PRICE_FLOOR = 21.95


def retail_price(wholesale, markup, floor=0.0):
    """wholesale x markup, rounded to the nearest .95 price point.

    `floor` lifts anything that would otherwise sell at a loss on a single-item
    order. Raising a $6.95 item to the floor is a real margin decision, not a
    rounding detail — see DEFAULT_PRICE_FLOOR.
    """
    target = wholesale * markup
    whole = int(target)
    # Candidates bracketing the target: X-1.95, X.95, X+0.95
    candidates = [whole - 1 + 0.95, whole + 0.95, whole + 1.95]
    candidates = [c for c in candidates if c >= 0.95]
    price = min(candidates, key=lambda c: abs(c - target))
    return max(price, floor)


def read_inventory():
    with INVENTORY.open(newline="", encoding="utf-8-sig") as fh:
        rows = list(csv.DictReader(fh))
    sellable = []
    for r in rows:
        if (r.get("DISCONTINUED") or "").strip().lower() == "yes":
            continue
        try:
            qty = int(r.get("QTY_AVAILABLE") or 0)
        except ValueError:
            qty = 0
        if qty <= 0:
            continue
        r["_qty"] = qty
        try:
            r["_wholesale"] = float(r.get("WHOLESALE_PRICE") or 0)
        except ValueError:
            r["_wholesale"] = 0.0
        sellable.append(r)
    return rows, sellable


def read_descriptions(catalogue, quiet=False):
    """Return {style: metadata} and {(style, color, size): row}.

    Metadata is style-level (copy, fabric, categories). Colour/size level rows
    carry the image filenames, which vary by colourway.
    """
    path = SOURCE / catalogue["workbook"]
    by_style = {}
    by_variant = {}
    for sheet, tag in catalogue["sheets"].items():
        header, rows = read_sheet(path, sheet, catalogue["images"], quiet=quiet)
        if header is None:
            print(f"  WARNING: sheet {sheet!r} not found in {path.name}")
            continue
        img_cols = image_columns(header)
        for raw in rows:
            row = dict(zip(header, raw))
            style = row.get("Style")
            if style is None or str(style).strip() == "":
                continue
            style = str(style).strip()
            row["_sheet"] = sheet
            row["_tag"] = tag
            row["_image_columns"] = img_cols
            by_style.setdefault(style, row)
            color = str(row.get("Color") or "").strip().upper()
            size = str(row.get("Size") or "").strip().upper()
            if size == "ONE SIZE":
                size = "O/S"
            by_variant[(style, color, size)] = row
    return by_style, by_variant


def clean(value):
    return str(value).strip() if value is not None else ""


def build_body_html(meta):
    desc = strip_trade_notes(clean(meta.get("Description")))
    fabric = clean(meta.get("Fabric/Material"))
    origin = clean(meta.get("Country Of Origin"))

    parts = []
    if desc:
        parts.append(f"<p>{desc.rstrip('. ')}.</p>")
    details = []
    if fabric:
        details.append(f"<li><strong>Fabric:</strong> {fabric}</li>")
    if origin:
        details.append(f"<li><strong>Made in:</strong> {origin}</li>")
    # Required by the Elegant Moments wholesale image/content licence.
    details.append(f"<li><strong>Brand:</strong> {ATTRIBUTION}</li>")
    parts.append("<ul>" + "".join(details) + "</ul>")
    # The workbook's "Shown With" column references other products by supplier
    # style number ("Shown with 1480."). That means nothing to a shopper and
    # cannot be clicked, so it is deliberately not rendered.
    return "".join(parts)


def build_tags(meta, styles):
    tags = []
    for key in ("Category 1", "Category 2", "Category 3"):
        val = clean(meta.get(key))
        if val:
            tags.append(val.rstrip(". "))
    sheet_tag = meta.get("_tag")
    if sheet_tag:
        tags.append(sheet_tag)
    # One tag per contributing supplier style, so a merged product still traces
    # back to both styles when placing a drop-ship order.
    for style in styles:
        tags.append(f"em-style-{style}")
    if len(styles) > 1:
        tags.append("Extended Sizing")
    seen, out = set(), []
    for t in tags:
        k = t.lower()
        if k not in seen:
            seen.add(k)
            out.append(t)
    return ", ".join(out)


def image_view_rank(name):
    """Front shots first, then back, then anything else.

    The workbook's "Image 1" column is usually the front view but not always —
    for a handful of styles it holds the back shot, which would otherwise become
    the featured image on collection tiles. Filenames follow a reliable
    `<style>_f` / `<style>_b` convention, so order on that instead.
    """
    lowered = name.lower()
    if "_f" in lowered:
        return 0
    if "_b" in lowered:
        return 1
    return 2


def image_names(row):
    names = []
    # Columns come from the sheet's own header, so a layout with Image 1..8
    # keeps its editorial shots instead of stopping at 4.
    for key in row.get("_image_columns") or ():
        val = clean(row.get(key))
        if val:
            names.append(val)
    # Stable sort keeps the supplier's ordering within each view type.
    return sorted(names, key=image_view_rank)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--catalogue", default=DEFAULT_CATALOGUE,
                    choices=list(CATALOGUES),
                    help=f"which descriptions workbook to build "
                         f"(default {DEFAULT_CATALOGUE})")
    ap.add_argument("--price-floor", type=float, default=DEFAULT_PRICE_FLOOR,
                    help="minimum retail price; below this a single-item order "
                         "loses money after the drop-ship fee, postage and "
                         "payment processing (default %(default)s, 0 disables)")
    ap.add_argument("--markup", type=float, default=2.75,
                    help="retail multiplier on wholesale cost (default 2.5)")
    ap.add_argument("--image-base", default="",
                    help="base URL for product images; when empty, Image Src is "
                         "left blank pending the catalog image manifest")
    ap.add_argument("--out", default=None,
                    help="output CSV path (default: the catalogue's out dir)")
    ap.add_argument("--status", default="draft", choices=["draft", "active"],
                    help="Shopify product status (default draft — review before "
                         "publishing)")
    ap.add_argument("--no-merge-plus", dest="merge_plus", action="store_false",
                    help="keep the supplier's split regular/plus styles as "
                         "separate products instead of merging them")
    ap.add_argument("--no-follows", dest="follows", action="store_false",
                    help="do not exclude styles already covered by earlier "
                         "catalogues")
    args = ap.parse_args()

    catalogue = CATALOGUES[args.catalogue]
    descriptions = SOURCE / catalogue["workbook"]

    for path in (INVENTORY, descriptions):
        if not path.exists():
            sys.exit(f"Missing source file: {path}\n"
                     f"Download it from b2b.elegantmoments.com and place it in {SOURCE}")

    all_rows, sellable = read_inventory()
    print(f"Catalogue: {args.catalogue} ({descriptions.name})")
    by_style, by_variant = read_descriptions(catalogue)

    # Workbooks overlap: a style carried over from an earlier release appears in
    # both, and importing it twice would collide on the em-<style> handle. The
    # first catalogue to ship a style owns it.
    already_imported = set()
    if args.follows:
        for earlier in catalogue["follows"]:
            prior, _ = read_descriptions(CATALOGUES[earlier], quiet=True)
            already_imported |= set(prior)
        overlap = set(by_style) & already_imported
        if overlap:
            print(f"  {len(overlap)} style(s) already covered by "
                  f"{', '.join(catalogue['follows'])} — left to that catalogue")

    # Group sellable inventory by style, keeping only styles we have copy for.
    styles = OrderedDict()
    skipped_no_copy = set()
    for r in sellable:
        style = clean(r.get("STYLE"))
        if style in already_imported:
            continue
        if style not in by_style:
            skipped_no_copy.add(style)
            continue
        styles.setdefault(style, []).append(r)

    # Elegant Moments splits sizing across two styles — 11022 (S–XL) and 11022X
    # (1X–4X) — which would otherwise list the same garment twice. Fold each plus
    # style into its parent as extra size variants. SKUs stay distinct, so
    # drop-ship order entry still identifies the correct supplier style.
    # The suffix is per-catalogue: 2026 uses X, Vivace uses Q for Queen.
    merged_styles = {}  # plus style -> parent style
    if args.merge_plus:
        suffixes = catalogue["plus_suffix"]
        for style in list(styles):
            if not style.upper().endswith(tuple(suffixes)):
                continue
            parent = style[:-1]
            if parent not in styles:
                continue
            plus_sizes = {clean(r.get("SIZE")).upper() for r in styles[style]}
            parent_sizes = {clean(r.get("SIZE")).upper() for r in styles[parent]}
            if plus_sizes & parent_sizes:
                # Overlapping sizes would collide as duplicate option pairs.
                continue
            styles[parent].extend(styles.pop(style))
            merged_styles[style] = parent

    image_base = args.image_base.rstrip("/") + "/" if args.image_base else ""
    out_path = Path(args.out) if args.out \
        else catalogue["out"] / "shopify_products.csv"
    out_path.parent.mkdir(parents=True, exist_ok=True)

    variant_count = 0
    # handle -> [image filename, ...]; consumed by push_products.py, which uploads
    # the bytes directly rather than relying on a public URL (we have none).
    manifest = OrderedDict()
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=SHOPIFY_COLUMNS)
        writer.writeheader()

        for style, rows in styles.items():
            meta = by_style[style]
            handle = f"em-{slugify(style)}"
            title = titleize(meta.get("Description"), style)

            colors = sorted({clean(r.get("COLOR")).upper() for r in rows})
            sizes = {clean(r.get("SIZE")).upper() for r in rows}
            multi_size = not (len(sizes) == 1 and next(iter(sizes)) in ("O/S", "OS"))

            rows.sort(key=lambda r: (clean(r.get("COLOR")).upper(),
                                     size_sort_key(clean(r.get("SIZE")))))

            # Contributing styles: the product's own, plus any merged plus twin.
            # Each ships its own photography, so gather images from both.
            contributing = [style] + [p for p, par in merged_styles.items()
                                      if par == style]

            # Collect images across each style's colourways, de-duplicated.
            imgs, seen_img = [], set()
            for src_style in contributing:
                src_meta = by_style.get(src_style, meta)
                for color in colors:
                    for size in sorted(sizes, key=size_sort_key):
                        vrow = by_variant.get((src_style, color, size)) or \
                               by_variant.get((src_style, color, "O/S"))
                        if vrow:
                            for name in image_names(vrow):
                                if name not in seen_img:
                                    seen_img.add(name)
                                    imgs.append(name)
                            break
                for name in image_names(src_meta):
                    if name not in seen_img:
                        seen_img.add(name)
                        imgs.append(name)

            if imgs:
                manifest[handle] = {"title": title, "images": imgs}

            weight_oz = clean(meta.get("Weight (oz)"))
            try:
                grams = int(round(float(weight_oz) * 28.3495)) if weight_oz else 0
            except ValueError:
                grams = 0

            for i, r in enumerate(rows):
                first = i == 0
                price = retail_price(r["_wholesale"], args.markup, args.price_floor)
                record = {c: "" for c in SHOPIFY_COLUMNS}
                record.update({
                    "Handle": handle,
                    "Option1 Name": "Color" if first else "",
                    "Option1 Value": clean(r.get("COLOR")).title(),
                    "Option2 Name": ("Size" if first and multi_size else ""),
                    "Option2 Value": (clean(r.get("SIZE")).upper() if multi_size else ""),
                    "Variant SKU": clean(r.get("SKU")),
                    "Variant Grams": grams,
                    "Variant Inventory Tracker": "shopify",
                    "Variant Inventory Qty": r["_qty"],
                    "Variant Inventory Policy": "deny",
                    "Variant Fulfillment Service": "manual",
                    "Variant Price": f"{price:.2f}",
                    "Variant Requires Shipping": "TRUE",
                    "Variant Taxable": "TRUE",
                    "Variant Barcode": clean(r.get("UPC")),
                })
                if first:
                    record.update({
                        "Title": f"{title} — Style {style}",
                        "Body (HTML)": build_body_html(meta),
                        "Vendor": ATTRIBUTION,
                        "Type": clean(meta.get("Category 1")).rstrip(". ") or "Lingerie",
                        "Tags": build_tags(meta, contributing),
                        "Published": "TRUE" if args.status == "active" else "FALSE",
                        "Status": args.status,
                    })
                    if imgs and image_base:
                        record["Image Src"] = image_base + imgs[0]
                        record["Image Position"] = 1
                        record["Image Alt Text"] = f"{title} by {ATTRIBUTION}"
                writer.writerow(record)
                variant_count += 1

            # Additional images ride on their own rows, keyed by handle.
            if image_base and len(imgs) > 1:
                for pos, name in enumerate(imgs[1:], start=2):
                    extra = {c: "" for c in SHOPIFY_COLUMNS}
                    extra.update({
                        "Handle": handle,
                        "Image Src": image_base + name,
                        "Image Position": pos,
                        "Image Alt Text": f"{title} by {ATTRIBUTION}",
                    })
                    writer.writerow(extra)

    manifest_path = out_path.parent / "image_manifest.json"
    with manifest_path.open("w", encoding="utf-8") as fh:
        json.dump(manifest, fh, indent=2)

    total_styles_seen = len({clean(r.get("STYLE")) for r in sellable})
    print(f"Inventory rows:        {len(all_rows)}")
    print(f"  sellable:            {len(sellable)}  ({total_styles_seen} styles)")
    print(f"Description styles:    {len(by_style)}")
    print(f"Products written:      {len(styles)}")
    if args.merge_plus:
        print(f"  plus styles merged:  {len(merged_styles)} "
              f"(folded into their regular-size parent)")
    print(f"Variant rows written:  {variant_count}")
    print(f"Styles without copy:   {len(skipped_no_copy)} (not in this "
          f"catalogue's workbook)")
    if not image_base:
        print("Images:                NONE — rerun with --image-base once the "
              "Catalog Images manifest is available")
    print(f"\nWrote {out_path}")
    print(f"Wrote {manifest_path} ({len(manifest)} products with images)")


if __name__ == "__main__":
    main()
